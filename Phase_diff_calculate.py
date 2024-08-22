import os
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from obspy.signal.detrend import polynomial

def Phase_diff(excel_file):
    # 創建結果文件夾
    output_folder = os.path.join(os.path.dirname(excel_file), 'image')
    os.makedirs(output_folder, exist_ok=True)

    # 創建結果文本文件
    results_file_path = os.path.join(os.path.dirname(excel_file), 'phase_diff_results.txt')
    with open(results_file_path, 'w') as results_file:
        results_file.write("============================\n")

        # 讀取 Excel 檔案
        df = pd.read_excel(excel_file)
        wb = openpyxl.load_workbook(excel_file)
        s1 = wb['Sheet1']
        ws = wb.active  

        # 去趨勢處理並繪圖
        s1['D1'].value = 'Ly_detrend'
        idxL = 2

        # 使用 polynomial 去趨勢處理並繪製圖形
        detrended_Ly = polynomial(df['L_Heel_29'], order=2, plot=True)
        x1 = detrended_Ly
        for element in x1:
            s1.cell(idxL, 4).value = element
            idxL += 1

        s1['E1'].value = 'Ry_detrend'
        idxR = 2

        # 使用 polynomial 去趨勢處理並繪製圖形
        detrended_Ry = polynomial(df['R_Heel_30'], order=2, plot=True)
        x2 = detrended_Ry
        for element in x2:
            s1.cell(idxR, 5).value = element
            idxR += 1

        # 刪除 F 到 I 列
        ws.delete_cols(6, 4) 

        # 儲存為新檔案名稱
        new_data_file_path = os.path.join(os.path.dirname(excel_file), 'EachFrame_detrended.xlsx')
        wb.save(new_data_file_path)

        # 讀取經過去趨勢處理後的數據
        df = pd.read_excel(new_data_file_path)

        # 繪製原始與去趨勢後的時間信號圖
        plt.figure(figsize=(10, 6))

        # 原始 Ly 信號
        plt.subplot(2, 1, 1)
        plt.plot(df['L_Heel_29'], label='Original Ly')
        plt.title('Original Ly Time Series')
        plt.xlabel('Sample Index')
        plt.ylabel('Amplitude')
        plt.legend()

        # 原始 Ry 信號
        plt.subplot(2, 1, 2)
        plt.plot(df['R_Heel_30'], label='Original Ry', color='orange')
        plt.title('Original Ry Time Series')
        plt.xlabel('Sample Index')
        plt.ylabel('Amplitude')
        plt.legend()

        plt.tight_layout()
        plt.savefig(os.path.join(output_folder, 'Ry_Ly_time_series.png'))
        plt.close()


        # plt.figure(figsize=(10, 6))

        # # 原始 Ly 信號
        # plt.subplot(2, 1, 1)
        # plt.plot(df['L_Heel_29'], label='Original Ly')
        # plt.title('Original Ly Time Series')
        # plt.xlabel('Sample Index')
        # plt.ylabel('Amplitude')
        # plt.legend()

        # # 去趨勢後的 Ly 信號
        # plt.subplot(2, 1, 2)
        # plt.plot(df['Ly_detrend'], label='Detrended Ly', color='orange')
        # plt.title('Detrended Ly Time Series')
        # plt.xlabel('Sample Index')
        # plt.ylabel('Amplitude')
        # plt.legend()

        # plt.tight_layout()
        # plt.savefig(os.path.join(output_folder, 'Ly_time_series.png'))
        # plt.close()


        # plt.figure(figsize=(10, 6))

        # # 原始 Ry 信號
        # plt.subplot(2, 1, 2)
        # # plt.subplot(2, 1, 1)
        # plt.plot(df['R_Heel_30'], label='Original Ry', color='orange')
        # plt.title('Original Ry Time Series')
        # plt.xlabel('Sample Index')
        # plt.ylabel('Amplitude')
        # plt.legend()

        # # 去趨勢後的 Ry 信號
        # plt.subplot(2, 1, 2)
        # plt.plot(df['Ry_detrend'], label='Detrended Ry', color='blue')
        # plt.title('Detrended Ry Time Series')
        # plt.xlabel('Sample Index')
        # plt.ylabel('Amplitude')
        # plt.legend()

        # plt.tight_layout()
        # plt.savefig(os.path.join(output_folder, 'Ry_time_series.png'))
        # plt.close()

        # ---------------[FFT]---------------------
        df = pd.read_excel(new_data_file_path)

        N = len(df['Ly_detrend'])
        results_file.write(f"Frame : {N}\n")
        results_file.write("============================\n\n")

        # 設置信號初始值及變數
        x = np.linspace(0, N/30, N)  # 每一點~=0.03s
        y1 = df['Ly_detrend']
        y2 = df['Ry_detrend']

        # 做FFT變換，以及抓取相位角
        ffted1 = np.fft.fft(y1)
        phase1 = np.angle(ffted1)

        ffted2 = np.fft.fft(y2)
        phase2 = np.angle(ffted2)

        # 計算出頻率軸的範圍
        freq = np.fft.fftfreq(len(x), x[1] - x[0])

        MaxL = 0
        idx = 0
        Max_idxL = 0
        for amplitude_e in np.abs(ffted1[0:N//2]):
            if amplitude_e > MaxL:
                MaxL = amplitude_e
                Max_idxL = idx
            idx += 1

        MaxR = 0
        idx = 0
        Max_idxR = 0
        for amplitude in np.abs(ffted2[0:N//2]):
            if amplitude > MaxR:
                MaxR = amplitude
                Max_idxR = idx
            idx += 1

        if Max_idxL != Max_idxR:
            results_file.write(f"LeftIndex : {Max_idxL}  {freq[Max_idxL]}   RightIndex : {Max_idxR}  {freq[Max_idxR]}\n")
        else:
            results_file.write("(主要頻率成分相位差)\n\n")
            results_file.write(f"主要頻率 : {freq[Max_idxL]}\n強度 : {MaxL}\n")
            results_file.write(f"主要頻率成分相位差 : {phase1[Max_idxL] - phase2[Max_idxL]}\n\n")

            results_file.write("============================\n\n")

            # 各頻率之相位差加總取平均
            difference = np.abs(phase2[0:N//2] - phase1[0:N//2])
            # 前100%強度的頻率
            arr = np.abs(ffted1[0:N//2])
            sum_amplitude = np.sum(arr)
            result_percent = arr / sum_amplitude

            sorted_indices = np.argsort(result_percent)
            result_indices = sorted_indices

            result_freq = []
            half_sum = 0
            for i in result_indices:
                result_freq.append(freq[i])
                half_sum += result_percent[i] * difference[i]

            half_average = half_sum / len(result_indices)
            results_file.write("(整體頻率成分加權相位差)\n\n")
            results_file.write(f"整體頻率相位差*相應百分比之加總 : {half_sum}\n")
            results_file.write(f"總和之平均 : {half_average}\n\n")
            results_file.write("============================")

        # 繪製FFT圖形
        plt.figure(figsize=(12, 8))

        # Left FFT
        plt.subplot(2, 1, 1)
        plt.plot(freq[0:N//2], np.abs(ffted1[0:N//2]), label='Left FFT', color='blue')
        plt.title('Left FFT')
        plt.xlabel('Frequency')
        plt.ylabel('Amplitude')
        plt.legend()

        # Right FFT
        plt.subplot(2, 1, 2)
        plt.plot(freq[0:N//2], np.abs(ffted2[0:N//2]), label='Right FFT', color='red')
        plt.title('Right FFT')
        plt.xlabel('Frequency')
        plt.ylabel('Amplitude')
        plt.legend()

        plt.tight_layout()
        #plt.savefig(os.path.join(output_folder, 'FFT_comparison.png'))
        plt.close()

        plt.figure()
        plt.xlabel('frequency') 
        plt.ylabel('Phase_difference')
        line1, = plt.plot(freq[0:N//2], (phase2[0:N//2] - phase1[0:N//2]), color='red', linewidth=1, label='Phase Difference')
        line2, = plt.plot(freq[0:N//2], (phase1[0:N//2] - phase1[0:N//2]), color='blue', linewidth=1, label='Baseline')
        plt.legend(handles=[line1, line2], loc='upper right')
        plt.tight_layout()
        plt.savefig(os.path.join(output_folder, 'phase_difference.png'))  
        plt.close()