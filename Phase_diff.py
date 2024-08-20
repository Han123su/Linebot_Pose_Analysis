import os
import pandas as pd
import matplotlib.pyplot as plt
from scipy.signal import detrend
import numpy as np
import openpyxl
from obspy.signal.detrend import polynomial

# 讀取Excel資料
PATH = "/data/2TSSD/Han212/skeleton/data/1016/1.25KG/"
data_file = "EachFrame.xlsx"
data_file_path = os.path.join(PATH, data_file)

# 新的檔案名稱
new_data_file = "EachFrame_detrended.xlsx"
new_data_file_path = os.path.join(PATH, new_data_file)


output_folder = os.path.join(PATH, 'image')
os.makedirs(output_folder, exist_ok=True)

df = pd.read_excel(data_file_path)
wb = openpyxl.load_workbook(data_file_path)
s1 = wb['Sheet1']
ws = wb.active  

# 去趨勢處理並繪圖
s1['D1'].value = 'Ly_detrend'
idxL = 2

# 使用polynomial去趨勢處理並繪製圖形
detrended_Ly = polynomial(df['L_Heel_29'], order=2, plot=True)
x1 = detrended_Ly
for element in x1:
    s1.cell(idxL, 4).value = element
    idxL = idxL + 1

s1['E1'].value = 'Ry_detrend'
idxR = 2

# 使用polynomial去趨勢處理並繪製圖形
detrended_Ry = polynomial(df['R_Heel_30'], order=2, plot=True)
x2 = detrended_Ry
for element in x2:
    s1.cell(idxR, 5).value = element
    idxR = idxR + 1

# 刪除 F 到 I 列
ws.delete_cols(6, 4) 

# 儲存為新檔案名稱
wb.save(new_data_file_path)

# 讀取經過去趨勢處理後的數據
df = pd.read_excel(new_data_file_path)

# 繪製原始與去趨勢後的時間信號圖
plt.figure(figsize=(10, 6))

# 原始 Ly 信號
plt.subplot(2, 1, 1)
plt.plot(df['L_Heel_29'], label='Original Ly')
plt.title('Original Ly Time Series')
plt.xlabel('Sample Index')
plt.ylabel('Amplitude')
plt.legend()

# 去趨勢後的 Ly 信號
plt.subplot(2, 1, 2)
plt.plot(df['Ly_detrend'], label='Detrended Ly', color='orange')
plt.title('Detrended Ly Time Series')
plt.xlabel('Sample Index')
plt.ylabel('Amplitude')
plt.legend()

plt.tight_layout()
plt.savefig(os.path.join(output_folder, 'Ly_time_series.png'))
plt.close()

# 繪製原始與去趨勢後的時間信號圖
plt.figure(figsize=(10, 6))

# 原始 Ry 信號
plt.subplot(2, 1, 1)
plt.plot(df['R_Heel_30'], label='Original Ry', color='orange')
plt.title('Original Ry Time Series')
plt.xlabel('Sample Index')
plt.ylabel('Amplitude')
plt.legend()

# 去趨勢後的 Ry 信號
plt.subplot(2, 1, 2)
plt.plot(df['Ry_detrend'], label='Detrended Ry', color='blue')
plt.title('Detrended Ry Time Series')
plt.xlabel('Sample Index')
plt.ylabel('Amplitude')
plt.legend()

plt.tight_layout()
plt.savefig(os.path.join(output_folder, 'Ry_time_series.png'))
plt.close()

# ---------------[FFT]---------------------
df = pd.read_excel(new_data_file_path)

N = len(df['Ly_detrend'])
print(N)

# 設置信號初始值及變數
x = np.linspace(0, N/30, N)  # 每一點~=0.03s
y1 = df['Ly_detrend']
y2 = df['Ry_detrend']

# 做FFT變換，以及抓取相位角
ffted1 = np.fft.fft(y1)
phase1 = np.angle(ffted1)

ffted2 = np.fft.fft(y2)
phase2 = np.angle(ffted2)

# 印出結果振福
print(phase1)
print("------------------------------------------")
print(phase2)

# 計算出頻率軸的範圍
print("------------------------------------------")
freq = np.fft.fftfreq(len(x), x[1] - x[0])
print(freq[0:N//2])

MaxL = 0
idx = 0
Max_idxL = 0
for amplitude_e in np.abs(ffted1[0:N//2]):
    if amplitude_e > MaxL:
        MaxL = amplitude_e
        Max_idxL = idx
    idx = idx + 1

MaxR = 0
idx = 0
Max_idxR = 0
for amplitude in np.abs(ffted2[0:N//2]):
    if amplitude > MaxR:
        MaxR = amplitude
        Max_idxR = idx
    idx = idx + 1

if(Max_idxL != Max_idxR):
    print("LeftIndex : " + str(Max_idxL) + "  " + str(freq[Max_idxL]) + "   RightIndex : " + str(Max_idxR) + "  " + str(freq[Max_idxR]))
else:
    print("Index : " + str(Max_idxL) + " amplitude : " + str(MaxL) + " freq : " + str(freq[Max_idxL]))
    print("PhaseLeft : " + str(phase1[Max_idxL]) + " PhaseRight : " + str(phase2[Max_idxL]))
    print("主要頻率成分相位差 : " + str((phase1[Max_idxL] - phase2[Max_idxL])))

    print('-------------------------------------')

    # 各頻率之相位差加總取平均
    difference = np.abs(phase2[0:N//2] - phase1[0:N//2])
    # 前100%強度的頻率
    arr = np.abs(ffted1[0:N//2])
    sum_amplitude = 0
    for e in arr:
        sum_amplitude = sum_amplitude + e
    result_percent = arr / sum_amplitude

    sorted_indices = np.argsort(result_percent)
    result_indices = sorted_indices

    result_freq = []
    half_sum = 0
    for i in result_indices:
        result_freq = np.append(result_freq, freq[i])
        half_sum = half_sum + result_percent[i] * difference[i]

    half_average = half_sum / len(result_indices)
    print("整體頻率相位差*相應百分比之加總 : " + str(half_sum))
    print("總和之平均 : " + str(half_average))
    print("freq : " + str(len(freq)))

# 將結果呈現出來
# 繪製FFT圖形
plt.figure(figsize=(12, 8))

# Left FFT
plt.subplot(2, 1, 1)
plt.plot(freq[0:N//2], np.abs(ffted1[0:N//2]), label='Left FFT', color='blue')
plt.title('Left FFT')
plt.xlabel('Frequency')
plt.ylabel('Amplitude')
plt.legend()

# Right FFT
plt.subplot(2, 1, 2)
plt.plot(freq[0:N//2], np.abs(ffted2[0:N//2]), label='Right FFT', color='red')
plt.title('Right FFT')
plt.xlabel('Frequency')
plt.ylabel('Amplitude')
plt.legend()

plt.tight_layout()
plt.savefig(os.path.join(output_folder, 'FFT_comparison.png'))
plt.close()

plt.figure()
plt.xlabel('frequency') 
plt.ylabel('Phase_difference')
line1, = plt.plot(freq[0:N//2], (phase2[0:N//2] - phase1[0:N//2]), color='red', linewidth=1, label='Phase Difference')
line2, = plt.plot(freq[0:N//2], (phase1[0:N//2] - phase1[0:N//2]), color='blue', linewidth=1, label='Baseline')
plt.legend(handles=[line1, line2], loc='upper right')
plt.tight_layout()
plt.savefig(os.path.join(output_folder, 'phase_difference.png'))  
plt.close()

# # 繪製 FFT 後的相位圖
# plt.figure(figsize=(10, 8))

# # 繪製y1 FFT後的相位
# plt.subplot(4, 1, 1)
# plt.plot(freq[0:500], np.angle(ffted1[0:500]))
# plt.title('Left_FFT')
# plt.xlabel('frequency')
# plt.ylabel('Phase')

# # 繪製y2 FFT後的相位
# plt.subplot(4, 1, 2)
# plt.stem(freq[0:500], np.angle(ffted2[0:500]))
# plt.title('Right_FFT')
# plt.xlabel('frequency')
# plt.ylabel('Phase')

# plt.subplot(4, 1, 3)
# plt.plot(freq[0:500], np.angle(ffted1[0:500]))
# plt.title('Left_FFT')
# plt.xlabel('frequency')
# plt.ylabel('Phase')

# # 繪製y2 FFT後的相位
# plt.subplot(4, 1, 4)
# plt.plot(freq[0:500], np.angle(ffted2[0:500]))
# plt.title('Right_FFT')
# plt.xlabel('frequency')
# plt.ylabel('Phase')

# plt.tight_layout()
# plt.savefig(os.path.join(output_folder, 'fft_phase.png'))







